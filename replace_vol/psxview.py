# Volatility
# Copyright (C) 2007-2013 Volatility Foundation
# Copyright (c) 2010, 2011, 2012 Michael Ligh <michael.ligh@mnin.org>
#
# This file is part of Volatility.
#
# Volatility is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# Volatility is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with Volatility.  If not, see <http://www.gnu.org/licenses/>.
#

import volatility.utils as utils
import volatility.obj as obj
import volatility.plugins.common as common
import volatility.win32.tasks as tasks
import volatility.plugins.modscan as modscan
import volatility.plugins.filescan as filescan
import volatility.plugins.overlays.windows.windows as windows
import volatility.plugins.gui.sessions as sessions
import volatility.plugins.gui.windowstations as windowstations
from volatility.renderers import TreeGrid
from volatility.renderers.basic import Address
import volatility.debug as debug
import volatility.plugins.addrspaces.standard as standard
import volatility.plugins.addrspaces.pmemaddressspace as pmemaddressspace

try:
    from openpyxl.workbook import Workbook
    from openpyxl.writer.excel import ExcelWriter
    from openpyxl.cell import get_column_letter
    from openpyxl.styles import Color, Fill, Style, PatternFill, Border, Side, Alignment, Protection, Font
    from openpyxl.cell import Cell
    from openpyxl import load_workbook
    has_openpyxl = True 
except ImportError:
    has_openpyxl = False


#--------------------------------------------------------------------------------
# object classes 
#--------------------------------------------------------------------------------

class _PSP_CID_TABLE(windows._HANDLE_TABLE): #pylint: disable-msg=W0212
    """Subclass the Windows handle table object for parsing PspCidTable"""

    def get_item(self, entry, handle_value = 0):

        p = obj.Object("address", entry.Object.v(), self.obj_vm)

        handle = obj.Object("_OBJECT_HEADER",
                offset = (p & ~7) -
                self.obj_vm.profile.get_obj_offset('_OBJECT_HEADER', 'Body'),
                vm = self.obj_vm)

        return handle

#--------------------------------------------------------------------------------
# profile modifications  
#--------------------------------------------------------------------------------

class MalwarePspCid(obj.ProfileModification):
    before = ['WindowsOverlay', 'WindowsVTypes']
    conditions = {'os': lambda x: x == 'windows'}

    def modification(self, profile):
        profile.vtypes.update({"_PSP_CID_TABLE" : profile.vtypes["_HANDLE_TABLE"]})
        profile.merge_overlay({"_KDDEBUGGER_DATA64" : [None,
            {'PspCidTable': [None,
                    ["pointer", ["pointer", ['_PSP_CID_TABLE']]]],
            }]})
        profile.object_classes.update({
            '_PSP_CID_TABLE': _PSP_CID_TABLE,
        })

#--------------------------------------------------------------------------------
# psxview plugin
#--------------------------------------------------------------------------------

class PsXview(common.AbstractWindowsCommand, sessions.SessionsMixin):
    "Find hidden processes with various process listings"

    def __init__(self, config, *args):
        common.AbstractWindowsCommand.__init__(self, config, *args)
        config.add_option("PHYSICAL-OFFSET", short_option = 'P', default = False,
                          help = "Physical Offset", action = "store_true")
        config.add_option("APPLY-RULES", short_option = 'R', default = False,
                          help = "Apply known good rules", action = "store_true")

    @staticmethod
    def get_file_offset(process):

        addr_space = process.obj_vm 
        address = process.obj_offset

        # we're already at the file layer (i.e. psscan on a raw memory image)
        if isinstance(addr_space, standard.FileAddressSpace) or isinstance(addr_space, pmemaddressspace.MemAddressSpace):
            return address

        paddr = addr_space.translate(address)
        offset = paddr

        addr_space = addr_space.base 
        while not (isinstance(addr_space, standard.FileAddressSpace) and isinstance(addr_space, pmemaddressspace.PMemAddressSpace)):
            offset = addr_space.translate(offset)
            # device memory addresses won't translate, so restore the original value 
            if offset == None:
                offset = paddr 
                break
            addr_space = addr_space.base 

        return offset

    def check_pslist(self, all_tasks):
        """Enumerate processes from PsActiveProcessHead"""
        return dict((PsXview.get_file_offset(p), p) for p in all_tasks)

    def check_psscan(self):
        """Enumerate processes with pool tag scanning"""
        return dict((PsXview.get_file_offset(p), p)
                    for p in filescan.PSScan(self._config).calculate())

    def check_thrdproc(self, _addr_space):
        """Enumerate processes indirectly by ETHREAD scanning"""
        ret = dict()

        for ethread in modscan.ThrdScan(self._config).calculate():
            if ethread.ExitTime != 0:
                continue
            # Bounce back to the threads owner 
            process = None
            if hasattr(ethread.Tcb, 'Process'):
                process = ethread.Tcb.Process.dereference_as('_EPROCESS')
            elif hasattr(ethread, 'ThreadsProcess'):
                process = ethread.ThreadsProcess.dereference()
            # Make sure the bounce succeeded 
            if (process and process.ExitTime == 0 and
                    process.UniqueProcessId > 0 and
                    process.UniqueProcessId < 65535):
                ret[PsXview.get_file_offset(process)] = process

        return ret

    def check_sessions(self, addr_space):
        """Enumerate processes from session structures"""
        
        ret = dict()
        for session in self.session_spaces(addr_space):
            for process in session.processes():
                ret[PsXview.get_file_offset(process)] = process
                
        return ret

    def check_desktop_thread(self, addr_space):
        """Enumerate processes from desktop threads"""
        
        ret = dict()
        for windowstation in windowstations.WndScan(self._config).calculate():
            for desktop in windowstation.desktops():
                for thread in desktop.threads():
                    process = thread.ppi.Process.dereference()
                    if process == None:
                        continue
                    ret[PsXview.get_file_offset(process)] = process
                    
        return ret

    def check_pspcid(self, addr_space):
        """Enumerate processes by walking the PspCidTable"""
        ret = dict()

        # Follow the pointers to the table base
        kdbg = tasks.get_kdbg(addr_space)
        PspCidTable = kdbg.PspCidTable.dereference().dereference()

        # Walk the handle table
        for handle in PspCidTable.handles():
            if handle.get_object_type() == "Process":
                process = handle.dereference_as("_EPROCESS")
                ret[PsXview.get_file_offset(process)] = process

        return ret

    def check_csrss_handles(self, all_tasks):
        """Enumerate processes using the csrss.exe handle table"""
        ret = dict()

        for p in all_tasks:
            if str(p.ImageFileName).lower() == "csrss.exe":
                # Gather the handles to process objects
                for handle in p.ObjectTable.handles():
                    if handle.get_object_type() == "Process":
                        process = handle.dereference_as("_EPROCESS")
                        ret[PsXview.get_file_offset(process)] = process

        return ret

    def calculate(self):
        if self._config.OUTPUT == "xlsx" and not has_openpyxl:
            debug.error("You must install OpenPyxl 2.1.2 for xlsx format:\n\thttps://pypi.python.org/pypi/openpyxl")
        elif self._config.OUTPUT == "xlsx" and not self._config.OUTPUT_FILE:
            debug.error("You must specify an output *.xlsx file!\n\t(Example: --output-file=OUTPUT.xlsx)")

        addr_space = utils.load_as(self._config)

        all_tasks = list(tasks.pslist(addr_space))

        ps_sources = {}
        # The keys are names of process sources. The values
        # are dictionaries whose keys are physical process 
        # offsets and the values are _EPROCESS objects. 
        ps_sources['pslist'] = self.check_pslist(all_tasks)
        ps_sources['psscan'] = self.check_psscan()
        ps_sources['thrdproc'] = self.check_thrdproc(addr_space)
        ps_sources['csrss'] = self.check_csrss_handles(all_tasks)
        ps_sources['pspcid'] = self.check_pspcid(addr_space)
        ps_sources['session'] = self.check_sessions(addr_space)
        ps_sources['deskthrd'] = self.check_desktop_thread(addr_space)

        # Build a list of offsets from all sources
        seen_offsets = []
        for source in ps_sources.values():
            for offset in source.keys():
                if offset not in seen_offsets:
                    seen_offsets.append(offset)
                    yield offset, source[offset], ps_sources

    def render_xlsx(self, outfd, data):
        BoldStyle = Style(font=Font(name='Calibri',
                 size=11,
                 bold=True,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FFFFFFFF'),
            fill=PatternFill(fill_type="solid",
                 start_color='FF000000',
                 end_color='FF000000'))
        RedStyle = Style(font=Font(name='Calibri',
                 size=11,
                 bold=False,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000'),
            border=Border(left=Side(border_style="thick",
                                color='FF000000'),
                      right=Side(border_style="thick",
                                 color='FF000000'),
                      top=Side(border_style="thick",
                               color='FF000000'),
                      bottom=Side(border_style="thick",
                                  color='FF000000'),
                      diagonal=Side(border_style="thick",
                                    color='FF000000'),
                      diagonal_direction=0,
                      outline=Side(border_style="thick",
                                   color='FF000000'),
                      vertical=Side(border_style="thick",
                                    color='FF000000'),
                      horizontal=Side(border_style="thick",
                                     color='FF000000')),
            fill=PatternFill(start_color = 'FFFF0000',
                    end_color = 'FFFF0000',
                    fill_type = 'solid'))
        GreenStyle = Style(font=Font(name='Calibri',
                 size=11,
                 bold=False,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000'),
            fill=PatternFill(start_color = "FF00FF00",
                    end_color = "FF00FF00",
                    fill_type = "solid"))

        wb = Workbook(optimized_write = True)
        ws = wb.create_sheet()
        ws.title = "Psxview Output"
        ws.append(["Offset (P)",
                  "Name",
                  "PID",
                  "pslist", 
                  "psscan", 
                  "thrdproc", 
                  "pspcid",
                  "csrss", 
                  "session", 
                  "deskthrd",
                  "Exit Time"])
        total = 1
        for offset, process, ps_sources in data:
            incsrss = ps_sources['csrss'].has_key(offset)
            insession = ps_sources['session'].has_key(offset)
            indesktop = ps_sources['deskthrd'].has_key(offset)
            inpspcid = ps_sources['pspcid'].has_key(offset)
            inpslist = ps_sources['pslist'].has_key(offset)
            inthread = ps_sources['thrdproc'].has_key(offset)

            if self._config.APPLY_RULES:
                if not incsrss:
                    if str(process.ImageFileName).lower() in ["system", "smss.exe", "csrss.exe"]:
                        incsrss = "Okay"
                    elif process.ExitTime > 0:
                        incsrss = "Okay"
                if not insession:
                    if str(process.ImageFileName).lower() in ["system", "smss.exe"]:
                        insession = "Okay"
                    elif process.ExitTime > 0:
                        insession = "Okay"
                if not indesktop:
                    if str(process.ImageFileName).lower() in ["system", "smss.exe"]:
                        indesktop = "Okay"
                    elif process.ExitTime > 0:
                        indesktop = "Okay"
                if not inpspcid:
                    if process.ExitTime > 0:
                        inpspcid = "Okay"
                if not inpslist:
                    if process.ExitTime > 0:
                        inpslist = "Okay"
                if not inthread:
                    if process.ExitTime > 0:
                        inthread = "Okay"

            ws.append([hex(offset),
                str(utils.remove_unprintable(str(process.ImageFileName)) or ""),
                str(process.UniqueProcessId),
                str(inpslist),
                str(ps_sources['psscan'].has_key(offset)),
                str(inthread),
                str(inpspcid),
                str(incsrss),
                str(insession),
                str(indesktop),
                str(process.ExitTime or '')])
            total += 1
        wb.save(filename = self._config.OUTPUT_FILE)

        wb = load_workbook(filename = self._config.OUTPUT_FILE)
        ws = wb.get_sheet_by_name(name = "Psxview Output")
        for col in xrange(1, 12):
            ws.cell("{0}{1}".format(get_column_letter(col), 1)).style = BoldStyle
        for row in xrange(2, total + 1):
            for col in xrange(4, 11):
                if ws.cell("{0}{1}".format(get_column_letter(col), row)).value == "False":
                    ws.cell("{0}{1}".format(get_column_letter(col), row)).style = RedStyle
                else:
                    ws.cell("{0}{1}".format(get_column_letter(col), row)).style = GreenStyle
        wb.save(filename = self._config.OUTPUT_FILE)

    def unified_output(self, data):
        return TreeGrid([("Offset(P)", Address),
                       ("Name", str),
                       ("PID", int),
                       ("pslist", str),
                       ("psscan", str),
                       ("thrdproc", str),
                       ("pspcid", str),
                       ("csrss", str),
                       ("session", str),
                       ("deskthrd", str),
                       ("ExitTime", str)],
                        self.generator(data))

    def generator(self, data):
        for offset, process, ps_sources in data:

            incsrss = ps_sources['csrss'].has_key(offset)
            insession = ps_sources['session'].has_key(offset)
            indesktop = ps_sources['deskthrd'].has_key(offset)
            inpspcid = ps_sources['pspcid'].has_key(offset)
            inpslist = ps_sources['pslist'].has_key(offset)
            inthread = ps_sources['thrdproc'].has_key(offset)

            if self._config.APPLY_RULES:
                if not incsrss:
                    if str(process.ImageFileName).lower() in ["system", "smss.exe", "csrss.exe"]:
                        incsrss = "Okay"
                    elif process.ExitTime > 0:
                        incsrss = "Okay"
                if not insession:
                    if str(process.ImageFileName).lower() in ["system", "smss.exe"]:
                        insession = "Okay"
                    elif process.ExitTime > 0:
                        insession = "Okay"
                if not indesktop:
                    if str(process.ImageFileName).lower() in ["system", "smss.exe"]:
                        indesktop = "Okay"
                    elif process.ExitTime > 0:
                        indesktop = "Okay"
                if not inpspcid:
                    if process.ExitTime > 0:
                        inpspcid = "Okay"
                if not inpslist:
                    if process.ExitTime > 0:
                        inpslist = "Okay"
                if not inthread:
                    if process.ExitTime > 0:
                        inthread = "Okay"

            yield (0, [
                Address(offset),
                str(process.ImageFileName),
                int(process.UniqueProcessId),
                str(inpslist),
                str(ps_sources['psscan'].has_key(offset)),
                str(inthread),
                str(inpspcid),
                str(incsrss),
                str(insession),
                str(indesktop),
                str(process.ExitTime or ''), 
                ])

    def render_text(self, outfd, data):

        self.table_header(outfd, [('Offset(P)', '[addrpad]'),
                                  ('Name', '<20'),
                                  ('PID', '>6'),
                                  ('pslist', '5'),
                                  ('psscan', '5'),
                                  ('thrdproc', '5'),
                                  ('pspcid', '5'),
                                  ('csrss', '5'),
                                  ('session', '5'),
                                  ('deskthrd', '5'),
                                  ('ExitTime', ""),
                                  ])

        for offset, process, ps_sources in data:

            incsrss = ps_sources['csrss'].has_key(offset)
            insession = ps_sources['session'].has_key(offset)
            indesktop = ps_sources['deskthrd'].has_key(offset)
            inpspcid = ps_sources['pspcid'].has_key(offset)
            inpslist = ps_sources['pslist'].has_key(offset)
            inthread = ps_sources['thrdproc'].has_key(offset)

            if self._config.APPLY_RULES:
                if not incsrss:
                    if str(process.ImageFileName).lower() in ["system", "smss.exe", "csrss.exe"]:
                        incsrss = "Okay"
                    elif process.ExitTime > 0:
                        incsrss = "Okay"
                if not insession:
                    if str(process.ImageFileName).lower() in ["system", "smss.exe"]:
                        insession = "Okay"
                    elif process.ExitTime > 0:
                        insession = "Okay"
                if not indesktop:
                    if str(process.ImageFileName).lower() in ["system", "smss.exe"]:
                        indesktop = "Okay"
                    elif process.ExitTime > 0:
                        indesktop = "Okay"
                if not inpspcid:
                    if process.ExitTime > 0:
                        inpspcid = "Okay"
                if not inpslist:
                    if process.ExitTime > 0:
                        inpslist = "Okay"
                if not inthread:
                    if process.ExitTime > 0:
                        inthread = "Okay"

            self.table_row(outfd,
                offset,
                process.ImageFileName,
                process.UniqueProcessId,
                str(inpslist),
                str(ps_sources['psscan'].has_key(offset)),
                str(inthread),
                str(inpspcid),
                str(incsrss),
                str(insession),
                str(indesktop),
                str(process.ExitTime or ''), 
                )

